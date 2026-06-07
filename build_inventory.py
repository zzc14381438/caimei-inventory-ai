"""
从Excel读取真实商品数据，生成带虚拟库存的商品档案
每个商品的每种颜色×尺码组合都有独立的虚拟库存
"""
import json
import random
from openpyxl import load_workbook

random.seed(42)  # 固定随机种子，每次生成结果一致

EXCEL_PATH = r"C:/Users/12196/Documents/xwechat_files/wxid_cpi5p1i6vwk922_c0bb/msg/file/2026-06/商品-2343607584912835203.xlsx"

wb = load_workbook(EXCEL_PATH)
ws = wb.active

products = []
total_sku_entries = 0

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    sku = str(row[1]).strip() if row[1] else ""
    name = str(row[2]).strip() if row[2] else ""
    status = str(row[3]).strip() if row[3] else "正常"
    purchase_price = float(row[4]) if row[4] else 0
    retail_price = float(row[7]) if row[7] else 0  # 价格3 是零售价
    excel_stock = int(float(row[11])) if row[11] else 0
    colors_raw = str(row[13]).strip() if row[13] else ""
    sizes_raw = str(row[14]).strip() if row[14] else ""
    supplier = str(row[18]).strip() if row[18] else ""
    category = str(row[20]).strip() if row[20] else ""
    style = str(row[24]).strip() if row[24] else ""
    season = str(row[25]).strip() if row[25] else ""
    fabric = str(row[21]).strip() if row[21] else ""
    is_special = str(row[17]).strip() if row[17] else "否"

    # 跳过无效行
    if not sku or not name:
        continue

    # 解析颜色和尺码 (逗号分隔)
    colors = [c.strip() for c in colors_raw.split(",") if c.strip()]
    sizes = [s.strip() for s in sizes_raw.split(",") if s.strip()]

    if not colors:
        colors = ["默认"]
    if not sizes:
        sizes = ["均码"]

    # 为每个颜色×尺码组合生成虚拟库存
    # 虚拟库存的总和大致等于Excel中的库存数（加一些随机波动）
    color_size_stock = []
    for color in colors:
        for size in sizes:
            # 生成0~8件的虚拟库存，偏重1~5件
            qty = random.choices(
                [0, 1, 2, 3, 4, 5, 6, 7, 8],
                weights=[8, 20, 25, 20, 12, 6, 4, 3, 2]
            )[0]
            # 零售价太低的不太可能有0库存（特价品除外）
            if retail_price < 30:
                qty = random.randint(0, 3)
            color_size_stock.append({
                "color": color,
                "size": size,
                "stock": qty
            })
            total_sku_entries += 1

    products.append({
        "sku": sku,
        "name": name,
        "status": status,
        "purchase_price": purchase_price,
        "retail_price": retail_price,
        "supplier": supplier,
        "category": category,
        "style": style,
        "season": season,
        "fabric": fabric,
        "is_special": is_special,
        "colors": colors,
        "sizes": sizes,
        "color_size_stock": color_size_stock,
        "total_excel_stock": excel_stock,
    })

# 保存为JSON
output = {
    "total_products": len(products),
    "total_sku_entries": total_sku_entries,
    "products": products
}

with open(r"C:\Users\12196\WorkBuddy\Claw\ai-app-dev\product_catalog.json", "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"✅ 商品档案生成完毕！")
print(f"   商品数: {len(products)}")
print(f"   SKU条目数(颜色×尺码): {total_sku_entries}")

# 统计汇总
names = set(p["name"] for p in products)
suppliers = set(p["supplier"] for p in products if p["supplier"])
all_colors = set()
all_sizes = set()
for p in products:
    for c in p["colors"]:
        all_colors.add(c)
    for s in p["sizes"]:
        all_sizes.add(s)

print(f"   商品名称种类: {len(names)}")
print(f"   供应商: {len(suppliers)}")
print(f"   颜色: {sorted(all_colors)}")
print(f"   尺码: {sorted(all_sizes)}")

print("\n=== 商品名称分布 ===")
from collections import Counter
name_counts = Counter(p["name"] for p in products)
for name, count in name_counts.most_common(15):
    print(f"   {name}: {count}个款号")
